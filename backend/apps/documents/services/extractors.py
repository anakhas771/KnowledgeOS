from pathlib import Path

import pymupdf
from docx import Document as DocxDocument
from openpyxl import load_workbook


class DocumentExtractionError(Exception):
    """Raised when document text extraction fails."""


def extract_text_from_txt(file_path: str) -> str:
    path = Path(file_path)

    try:
        return path.read_text(encoding="utf-8")
    except UnicodeDecodeError as exc:
        raise DocumentExtractionError(
            "Text file is not valid UTF-8."
        ) from exc


def extract_text_from_pdf(file_path: str) -> str:
    try:
        document = pymupdf.open(file_path)

        pages = [
            page.get_text("text")
            for page in document
        ]

        document.close()

        return "\n".join(pages)

    except Exception as exc:
        raise DocumentExtractionError(
            "Failed to extract text from PDF."
        ) from exc


def extract_text_from_docx(file_path: str) -> str:
    try:
        document = DocxDocument(file_path)

        paragraphs = [
            paragraph.text
            for paragraph in document.paragraphs
            if paragraph.text.strip()
        ]

        return "\n".join(paragraphs)

    except Exception as exc:
        raise DocumentExtractionError(
            "Failed to extract text from DOCX."
        ) from exc


def extract_text_from_xlsx(file_path: str) -> str:
    try:
        workbook = load_workbook(
            filename=file_path,
            read_only=True,
            data_only=True,
        )

        sections: list[str] = []

        for worksheet in workbook.worksheets:
            sections.append(f"[Sheet: {worksheet.title}]")

            for row in worksheet.iter_rows(values_only=True):
                values = [
                    str(value).strip()
                    for value in row
                    if value is not None
                ]

                if values:
                    sections.append(" | ".join(values))

        workbook.close()

        return "\n".join(sections)

    except Exception as exc:
        raise DocumentExtractionError(
            "Failed to extract text from XLSX."
        ) from exc


def extract_text(
    file_path: str,
    file_type: str | None = None,
) -> str:
    """
    Extract textual content based on extension or MIME type.
    """

    path = Path(file_path)

    extension = path.suffix.lower()

    if extension in {".txt", ".md"}:
        return extract_text_from_txt(file_path)

    if extension == ".pdf":
        return extract_text_from_pdf(file_path)

    if extension == ".docx":
        return extract_text_from_docx(file_path)

    if extension == ".xlsx":
        return extract_text_from_xlsx(file_path)

    if file_type == "application/pdf":
        return extract_text_from_pdf(file_path)

    if file_type in {
        "text/plain",
        "text/markdown",
    }:
        return extract_text_from_txt(file_path)

    raise DocumentExtractionError(
        f"Unsupported document type: {extension or file_type}"
    )
